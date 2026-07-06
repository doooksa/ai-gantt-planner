import pytest

from datetime import timedelta

from app.excel import export_excel, parse_excel
from app.domain.models import Op, Patch, Selector
from app.domain.patches import apply_patch
from app.domain.scheduler import schedule
from app.domain.validators import PlanValidationError
from app.storage.seed import seed_plan

from conftest import make_xlsx


def test_cyrillic_headers_with_extra_spaces_and_case():
    # Headers deliberately messy: extra spaces, mixed case, trailing/leading ws.
    headers = ["  Задача ", "ОПИСАНИЕ", " Исполнитель", "Длительность ", "  предшественники  "]
    rows = [
        ["Research", "Сбор требований", "Anna", 2, ""],
        ["Design", "Макеты", "Anna", 3, "Research"],
    ]
    plan = parse_excel(make_xlsx(headers, rows))

    assert [t.name for t in plan.tasks] == ["Research", "Design"]
    design = plan.by_id("design")
    research = plan.by_id("research")
    assert design.predecessor_ids == [research.id]


def test_empty_predecessors_gives_no_deps():
    headers = ["задача", "описание", "исполнитель", "длительность", "предшественники"]
    rows = [["Solo", "", "Ann", 4, ""]]
    plan = parse_excel(make_xlsx(headers, rows))
    assert plan.tasks[0].predecessor_ids == []


def test_multiple_predecessors_comma_separated():
    headers = ["задача", "длительность", "предшественники"]
    rows = [
        ["A", 1, ""],
        ["B", 1, ""],
        ["C", 1, "A, B"],
    ]
    plan = parse_excel(make_xlsx(headers, rows))
    c = plan.by_id("c")
    assert set(c.predecessor_ids) == {"a", "b"}


def test_invalid_duration_reports_row_number():
    headers = ["задача", "длительность", "предшественники"]
    rows = [
        ["A", 2, ""],
        ["B", "нет", ""],  # row 3 in the sheet (header=1, A=2, B=3)
    ]
    with pytest.raises(PlanValidationError) as exc:
        parse_excel(make_xlsx(headers, rows))
    assert exc.value.code == "bad_duration"
    assert "3" in exc.value.message


def test_unknown_predecessor_name_errors():
    headers = ["задача", "длительность", "предшественники"]
    rows = [["A", 1, "Ghost"]]
    with pytest.raises(PlanValidationError) as exc:
        parse_excel(make_xlsx(headers, rows))
    assert exc.value.code == "missing_predecessor"


def test_missing_required_column():
    headers = ["описание", "длительность"]  # no "задача"
    rows = [["x", 1]]
    with pytest.raises(PlanValidationError) as exc:
        parse_excel(make_xlsx(headers, rows))
    assert exc.value.code == "missing_column"


def test_blank_trailing_rows_are_skipped():
    headers = ["задача", "длительность", "предшественники"]
    rows = [
        ["A", 1, ""],
        [None, None, None],
        ["", "", ""],
    ]
    plan = parse_excel(make_xlsx(headers, rows))
    assert [t.name for t in plan.tasks] == ["A"]


def test_cycle_in_excel_rejected():
    headers = ["задача", "длительность", "предшественники"]
    rows = [
        ["A", 1, "B"],
        ["B", 1, "A"],
    ]
    with pytest.raises(PlanValidationError) as exc:
        parse_excel(make_xlsx(headers, rows))
    assert exc.value.code == "cycle"


def test_export_then_import_roundtrip(start):
    plan = seed_plan()
    data = export_excel(plan, start)
    reparsed = parse_excel(data)

    assert [t.name for t in reparsed.tasks] == [t.name for t in plan.tasks]
    # Dependencies survive the round-trip (resolved by name -> id).
    orig_name = {t.id: t.name for t in plan.tasks}
    for orig, back in zip(plan.tasks, reparsed.tasks):
        orig_pred_names = {orig_name[p] for p in orig.predecessor_ids}
        back_name = {t.id: t.name for t in reparsed.tasks}
        back_pred_names = {back_name[p] for p in back.predecessor_ids}
        assert orig_pred_names == back_pred_names


def test_export_import_drops_offset_days(start):
    """Conscious trade-off: a manual shift (offset_days) does not survive the
    Excel round-trip, because the format has no column for it and import ignores
    the computed date columns. Structure is preserved; the shift is not."""
    plan = seed_plan()
    shifted, _ = apply_patch(
        plan,
        Patch(ops=[Op(type="shift_task", selector=Selector(by_name="Frontend"), payload={"days": 3})]),
        start,
    )
    assert shifted.by_id("frontend").offset_days == 3

    # Exported dates DO reflect the shift...
    base = {t.id: t for t in schedule(plan.tasks, start)}
    shifted_sched = {t.id: t for t in schedule(shifted.tasks, start)}
    assert shifted_sched["frontend"].start == base["frontend"].start + timedelta(days=3)

    # ...but a round-trip resets offset_days to 0 (documented limitation).
    reparsed = parse_excel(export_excel(shifted, start))
    assert reparsed.by_id("frontend").offset_days == 0
    # Structure (tasks, deps, durations) is fully preserved.
    assert [t.name for t in reparsed.tasks] == [t.name for t in shifted.tasks]
    reparsed_sched = {t.id: t for t in schedule(reparsed.tasks, start)}
    assert reparsed_sched["frontend"].start == base["frontend"].start  # shift gone


def test_export_includes_computed_dates(start):
    from openpyxl import load_workbook
    import io

    plan = seed_plan()
    data = export_excel(plan, start)
    ws = load_workbook(io.BytesIO(data)).worksheets[0]
    header = [c.value for c in ws[1]]
    assert "дата начала" in header
    assert "дата конца" in header
