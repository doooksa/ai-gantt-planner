"""Excel import/export (openpyxl).

Import columns (case-insensitive, whitespace-normalized headers):
    задача, описание, исполнитель, длительность, предшественники
Only the FIRST sheet is read, with read_only=True.
Predecessors are task *names*, comma-separated, resolved to ids.
Export adds computed `дата начала`, `дата конца`.

All user-facing errors are Russian and include the row number where relevant.

Known limitation — `offset_days` is NOT preserved across an Excel round-trip.
    The interchange format is deliberately fixed at 5 structural columns
    (задача/описание/исполнитель/длительность/предшественники) + 2 *computed*
    date columns. A task's manual shift (`offset_days`, set via `shift_task`) is
    not one of those columns, so export→import resets it to 0. The exported date
    columns DO reflect the shift (they are scheduled with the offset applied),
    but on import those dates are ignored — dates are derived, never a source of
    truth — so the reimported plan is scheduled without the shift. This is a
    conscious trade-off to keep the format spec-exact and the "dates are derived"
    invariant intact; it is asserted by test_export_import_drops_offset_days and
    noted for the production roadmap (a hidden "сдвиг, дн." column would restore
    full fidelity if ever needed).
"""

from __future__ import annotations

import io
import re
from datetime import date
from typing import IO

from openpyxl import Workbook, load_workbook

from .domain.models import Plan, ScheduledPlan, Task
from .domain.scheduler import schedule
from .domain.slug import unique_slug
from .domain.validators import PlanValidationError

# Canonical header -> model field.
_HEADER_TO_FIELD = {
    "задача": "name",
    "описание": "description",
    "исполнитель": "assignee",
    "длительность": "duration_days",
    "предшественники": "predecessors",
}

_EXPORT_HEADERS = [
    "задача",
    "описание",
    "исполнитель",
    "длительность",
    "предшественники",
    "дата начала",
    "дата конца",
]


def _norm_header(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def _split_predecessors(value) -> list[str]:
    if value is None:
        return []
    parts = re.split(r"[,;]", str(value))
    return [p.strip() for p in parts if p.strip()]


def parse_excel(source: IO[bytes] | str | bytes) -> Plan:
    """Parse the first sheet of an .xlsx into a validated Plan.

    Raises PlanValidationError with a Russian message (and row number where
    applicable) on any problem.
    """
    if isinstance(source, bytes):
        source = io.BytesIO(source)
    try:
        wb = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises various types for bad files
        raise PlanValidationError(
            "Не удалось прочитать файл Excel. Ожидается .xlsx.",
            code="bad_file",
            detail=f"openpyxl load failed: {exc}",
        ) from exc

    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        raise PlanValidationError(
            "Файл пустой: нет строки заголовков.",
            code="empty_file",
            detail="no header row",
        )

    # Map normalized header -> column index.
    col_of_field: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        field = _HEADER_TO_FIELD.get(_norm_header(cell))
        if field and field not in col_of_field:
            col_of_field[field] = idx

    if "name" not in col_of_field:
        raise PlanValidationError(
            "В таблице отсутствует обязательный столбец «задача».",
            code="missing_column",
            detail=f"headers seen: {[_norm_header(c) for c in header_row]}",
        )

    # First pass: build tasks with names + pending predecessor names.
    tasks: list[Task] = []
    pending_preds: list[list[str]] = []
    used_ids: set[str] = set()

    for row_number, row in enumerate(rows, start=2):
        def cell(field: str):
            i = col_of_field.get(field)
            return row[i] if i is not None and i < len(row) else None

        name = cell("name")
        if name is None or str(name).strip() == "":
            # Skip fully blank rows (trailing empties are common in xlsx).
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            raise PlanValidationError(
                f"Строка {row_number}: не заполнено название задачи.",
                code="missing_name",
                detail=f"row {row_number} without name",
            )
        name = str(name).strip()

        duration_raw = cell("duration_days")
        try:
            duration = int(duration_raw)
        except (TypeError, ValueError):
            raise PlanValidationError(
                f"Строка {row_number}: длительность «{duration_raw}» "
                f"должна быть целым числом.",
                code="bad_duration",
                detail=f"row {row_number} duration={duration_raw!r}",
            )
        if duration < 1:
            raise PlanValidationError(
                f"Строка {row_number}: длительность должна быть не меньше 1.",
                code="bad_duration",
                detail=f"row {row_number} duration={duration}",
            )

        description = cell("description")
        assignee = cell("assignee")
        task_id = unique_slug(name, used_ids, index=len(tasks))
        used_ids.add(task_id)

        tasks.append(
            Task(
                id=task_id,
                name=name,
                description=str(description).strip() if description not in (None, "") else None,
                assignee=str(assignee).strip() if assignee not in (None, "") else None,
                duration_days=duration,
            )
        )
        pending_preds.append(_split_predecessors(cell("predecessors")))

    wb.close()

    # Second pass: resolve predecessor names -> ids.
    name_to_id: dict[str, str] = {}
    for t in tasks:
        name_to_id.setdefault(_norm_header(t.name), t.id)

    for t, preds in zip(tasks, pending_preds):
        resolved: list[str] = []
        for pred_name in preds:
            pid = name_to_id.get(_norm_header(pred_name))
            if pid is None:
                raise PlanValidationError(
                    f"Задача «{t.name}»: предшественник «{pred_name}» "
                    f"не найден среди задач таблицы.",
                    code="missing_predecessor",
                    detail=f"{t.id} -> unknown predecessor name {pred_name!r}",
                )
            if pid != t.id and pid not in resolved:
                resolved.append(pid)
        t.predecessor_ids = resolved

    # Final structural validation (cycles, etc.) via the scheduler.
    schedule(tasks)
    return Plan(version=0, tasks=tasks)


def export_excel(plan: Plan, project_start: date | None = None) -> bytes:
    """Serialize a plan (with computed dates) to .xlsx bytes."""
    scheduled = schedule(plan.tasks, project_start)
    id_to_name = {t.id: t.name for t in plan.tasks}

    wb = Workbook()
    ws = wb.active
    ws.title = "План"
    ws.append(_EXPORT_HEADERS)

    for t in scheduled:
        preds = ", ".join(id_to_name.get(p, p) for p in t.predecessor_ids)
        ws.append([
            t.name,
            t.description or "",
            t.assignee or "",
            t.duration_days,
            preds,
            t.start.isoformat(),
            t.end.isoformat(),
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def scheduled_plan(plan: Plan, project_start: date | None = None) -> ScheduledPlan:
    """Convenience: full scheduled view of a plan."""
    if project_start is None:
        project_start = date.today()
    return ScheduledPlan(
        version=plan.version,
        project_start=project_start,
        tasks=schedule(plan.tasks, project_start),
    )
